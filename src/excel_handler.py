"""
OpenClaw Excel Handler - 精简版
专为OpenClaw优化，去除冗余功能，保留核心能力
"""

from pathlib import Path
from typing import Union, Optional, Dict
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment


class ExcelHandler:
    """OpenClaw专用的Excel文件处理器"""
    
    def __init__(self):
        """初始化处理器"""
        self.file_path = None
    
    def read_excel(self, file_path: Union[str, Path], 
                   sheet_name: Union[str, int, None] = None) -> pd.DataFrame:
        """
        读取Excel文件到DataFrame
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称（默认第一个）
            
        Returns:
            DataFrame
        """
        try:
            return pd.read_excel(file_path, sheet_name=sheet_name)
        except Exception as e:
            raise ValueError(f"Excel读取失败: {str(e)}")
    
    def write_excel(self, data: pd.DataFrame, 
                    file_path: Union[str, Path],
                    sheet_name: str = "Sheet1") -> None:
        """
        写入DataFrame到Excel文件
        
        Args:
            data: 要写入的DataFrame
            file_path: 输出文件路径
            sheet_name: 工作表名称
        """
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                data.to_excel(writer, sheet_name=sheet_name, index=False)
            self._style_worksheet(file_path, sheet_name)
        except Exception as e:
            raise ValueError(f"Excel写入失败: {str(e)}")
    
    def _style_worksheet(self, file_path: Union[str, Path], sheet_name: str):
        """为工作表应用OpenClaw标准样式"""
        try:
            wb = load_workbook(file_path)
            ws = wb[sheet_name]
            
            # 表头样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 自动列宽
            for column in ws.columns:
                max_length = 0
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                column_letter = get_column_letter(column[0].column)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(file_path)
        except Exception:
            pass  # 样式失败不影响数据
    
    def get_sheet_names(self, file_path: Union[str, Path]) -> list:
        """获取Excel文件的工作表名称列表"""
        try:
            wb = load_workbook(file_path)
            return wb.sheetnames
        except Exception as e:
            raise ValueError(f"无法读取工作表: {str(e)}")
    
    def filter_data(self, file_path: Union[str, Path], 
                   filters: Dict[str, any]) -> pd.DataFrame:
        """
        读取并过滤Excel数据
        
        Args:
            file_path: Excel文件路径
            filters: 过滤条件字典 {列名: 值}
            
        Returns:
            过滤后的DataFrame
        """
        df = self.read_excel(file_path)
        
        if filters:
            for column, value in filters.items():
                if column in df.columns:
                    df = df[df[column] == value]
        
        return df